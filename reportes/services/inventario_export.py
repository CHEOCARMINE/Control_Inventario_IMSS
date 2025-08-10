from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib.staticfiles import finders
from openpyxl.drawing.image import Image as XLImage
from django.db.models import Exists, OuterRef, Prefetch
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Columnas
COLUMNAS = [
    "ID","Nombre","Tipo de producto","Categoría","Marca","Modelo","Color",
    "Rol (Normal/Padre/Hijo)","Padre","Número de serie","Unidad",
    "Stock","Stock mínimo","Estado"
]

# Rutas a tus logos
LOGO_IZQ_STATIC = "salidas/img/logo_gob.png"
LOGO_DER_STATIC = "salidas/img/logo_imss.png"

# Estilos
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_HEADER = PatternFill("solid", fgColor="EDEDED")

# Tamaños del encabezado
IMG_MAX_HEIGHT_PX = 110
COL_LOGO_WIDTH = 32
HEADER_ROW_HEIGHT = 30

def _insertar_imagen(ws, static_path, cell):
    path = finders.find(static_path)
    if not path:
        return
    img = XLImage(path)
    # Escalar manteniendo proporción a la altura objetivo
    try:
        ratio = IMG_MAX_HEIGHT_PX / float(img.height)
        img.height = int(IMG_MAX_HEIGHT_PX)
        img.width = int(img.width * ratio)
    except Exception:
        pass
    img.anchor = cell
    ws.add_image(img)

def _encabezado_en_hoja(ws, texto_centro, total_cols):
    # Altura generosa de las 3 filas del encabezado
    for r in (1, 2, 3):
        ws.row_dimensions[r].height = HEADER_ROW_HEIGHT

    # Reservar columnas para logos (A y última) y ensancharlas
    last_letter = get_column_letter(total_cols)
    ws.column_dimensions["A"].width = COL_LOGO_WIDTH
    ws.column_dimensions[last_letter].width = COL_LOGO_WIDTH

    # Texto centrado: C1 : (última-2)3 para no chocar con los logos
    start_col = 3  # C
    end_col = max(4, total_cols - 2)
    ws.merge_cells(start_row=1, start_column=start_col, end_row=3, end_column=end_col)
    c = ws.cell(row=1, column=start_col)
    c.value = texto_centro
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Insertar logos dentro del área del header
    _insertar_imagen(ws, LOGO_IZQ_STATIC, "A1")
    _insertar_imagen(ws, LOGO_DER_STATIC, f"{last_letter}1")

def _titulos_tabla(ws, fila_titulo=5):
    for idx, nombre in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=fila_titulo, column=idx, value=nombre)
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Congelar encabezado + fila de títulos
    ws.freeze_panes = "A6"

    # AutoFilter
    last_col = get_column_letter(len(COLUMNAS))
    ws.auto_filter.ref = f"A{fila_titulo}:{last_col}{fila_titulo}"

    # Anchos sugeridos 
    widths = [8,34,18,18,16,16,12,18,22,22,10,12,12,12]
    for i, w in enumerate(widths, start=1):
        if i == 1 or i == len(COLUMNAS):
            continue
        ws.column_dimensions[get_column_letter(i)].width = w

def _crear_hoja(wb, nombre, texto_encabezado):
    ws = wb.create_sheet(title=nombre)
    _encabezado_en_hoja(ws, texto_encabezado, total_cols=len(COLUMNAS))
    _titulos_tabla(ws, fila_titulo=5)
    return ws

def crear_libro_base(texto_encabezado: str) -> Workbook:
    wb = Workbook()

    # Hoja inicial -> Activos
    ws0 = wb.active
    ws0.title = "Activos"
    _encabezado_en_hoja(ws0, texto_encabezado, total_cols=len(COLUMNAS))
    _titulos_tabla(ws0, fila_titulo=5)

    _crear_hoja(wb, "Normales", texto_encabezado)
    _crear_hoja(wb, "Padres e Hijos", texto_encabezado)
    _crear_hoja(wb, "Desactivados", texto_encabezado)

    return wb

# INDICES de columna por nombre (útil para fórmulas)
IDX = {name: i+1 for i, name in enumerate(COLUMNAS)}  

# Query base con relaciones (evita N+1)
def _qs_base(Producto):
    return (Producto.objects
        .select_related(
            "tipo__Subcatalogo__catalogo",
            "tipo__unidad_medida",
            "marca",
            "producto_padre",
        )
        .annotate(tiene_hijos=Exists(Producto.objects.filter(producto_padre_id=OuterRef("pk"))))
    )

# Helpers para mapear el producto a una fila
def _rol(p):
    if p.producto_padre_id:
        return "Hijo"
    return "Padre" if getattr(p, "tiene_hijos", False) else "Normal"

def _categoria(p):
    try:  
        return str(p.tipo.Subcatalogo.catalogo)
    except Exception:
        return ""

def _unidad(p):
    try:
        return str(p.tipo.unidad_medida)
    except Exception:
        return ""

def _padre_nombre(p):
    return str(p.producto_padre) if p.producto_padre_id else ""

def _stock_min(p):
    try:
        return int(p.tipo.stock_minimo)
    except Exception:
        return None

def fila_from_producto(p):
    return [
        p.id,
        p.nombre,
        str(p.tipo.nombre) if p.tipo_id else "",
        _categoria(p),
        str(p.marca) if p.marca_id else "",
        p.modelo,
        p.color,
        _rol(p),
        _padre_nombre(p),
        p.numero_serie or "",
        _unidad(p),
        int(p.stock or 0),
        _stock_min(p),
        "Activo" if p.estado else "Desactivado",
    ]

def _escribir_filas(ws, filas, fila_inicio=6):
    r = fila_inicio
    for row in filas:
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = BORDER
            if c_idx in (IDX["Stock"], IDX["Stock mínimo"]):
                cell.alignment = Alignment(horizontal="center")
        r += 1
    # Extiende AutoFilter al rango de datos
    last_col = get_column_letter(len(COLUMNAS))
    ws.auto_filter.ref = f"A5:{last_col}{max(5, r-1)}"
    return r - 1  

# Formato condicional: Stock (Activos/Normales: todos)
def _pintar_stock_general(ws, fila_inicio=6):
    colS = get_column_letter(IDX["Stock"])
    colM = get_column_letter(IDX["Stock mínimo"])
    last = ws.max_row
    if last < fila_inicio:
        return
    rango = f"{colS}{fila_inicio}:{colS}{last}"
    rojo = PatternFill(fill_type="solid", start_color="FFF8CBAD", end_color="FFF8CBAD")
    amarillo = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")
    ws.conditional_formatting.add(
        rango, CellIsRule(operator="equal", formula=["0"], stopIfTrue=True, fill=rojo)
    )
    formula = f'AND({colS}{fila_inicio}>0,{colS}{fila_inicio}<={colM}{fila_inicio})'
    ws.conditional_formatting.add(
        rango, FormulaRule(formula=[formula], stopIfTrue=True, fill=amarillo)
    )

def _pintar_stock_solo_padres(ws, fila_inicio=6):
    colS  = get_column_letter(IDX["Stock"])
    colM  = get_column_letter(IDX["Stock mínimo"])
    colR  = get_column_letter(IDX["Rol (Normal/Padre/Hijo)"])
    last  = ws.max_row
    if last < fila_inicio:
        return

    rango = f"{colS}{fila_inicio}:{colS}{last}"
    rojo = PatternFill(fill_type="solid", start_color="FFF8CBAD", end_color="FFF8CBAD")
    amarillo = PatternFill(fill_type="solid", start_color="FFFFF2CC", end_color="FFFFF2CC")

    formula_rojo = f'AND({colR}{fila_inicio}="Padre",{colS}{fila_inicio}=0)'
    ws.conditional_formatting.add(rango, FormulaRule(formula=[formula_rojo], stopIfTrue=True, fill=rojo))

    formula_amarillo = f'AND({colR}{fila_inicio}="Padre",{colS}{fila_inicio}>0,{colS}{fila_inicio}<={colM}{fila_inicio})'
    ws.conditional_formatting.add(rango, FormulaRule(formula=[formula_amarillo], stopIfTrue=True, fill=amarillo))

# Poblar hoja Activos
def poblar_activos(wb, Producto, texto_encabezado):
    ws = wb["Activos"] if "Activos" in wb.sheetnames else wb.active
    qs = _qs_base(Producto).filter(estado=True)
    filas = (fila_from_producto(p) for p in qs.iterator(chunk_size=1000))
    _escribir_filas(ws, filas)
    _pintar_stock_general(ws)

def poblar_normales(wb, Producto, _texto_encabezado=None):
    ws = wb["Normales"]
    qs = _qs_base(Producto).filter(
        estado=True,
        producto_padre__isnull=True,
        tiene_hijos=False
    )
    filas = (fila_from_producto(p) for p in qs.iterator(chunk_size=1000))
    _escribir_filas(ws, filas)
    _pintar_stock_general(ws)

def poblar_padres_hijos(wb, Producto, _texto_encabezado=None):
    ws = wb["Padres e Hijos"]

    # Prefetch de hijos activos con todas las relaciones necesarias
    hijos_qs = (Producto.objects
                .filter(estado=True)
                .select_related("tipo__Subcatalogo__catalogo", "tipo__unidad_medida", "marca", "producto_padre"))

    padres_qs = (_qs_base(Producto)
                .filter(estado=True, producto_padre__isnull=True, tiene_hijos=True)
                .prefetch_related(Prefetch("productos_hijos", queryset=hijos_qs)))

    def filas_generator():
        for padre in padres_qs.iterator(chunk_size=500):
            yield fila_from_producto(padre)
            for h in padre.productos_hijos.all():
                fila = fila_from_producto(h)
                fila[1] = f"— {fila[1]}"  
                yield fila

    _escribir_filas(ws, filas_generator())
    _pintar_stock_solo_padres(ws)

def poblar_desactivados(wb, Producto, _texto_encabezado=None):
    ws = wb["Desactivados"]
    qs = (_qs_base(Producto)
            .filter(estado=False)
            .order_by("tipo__Subcatalogo__catalogo__nombre", "tipo__nombre", "nombre"))
    filas = (fila_from_producto(p) for p in qs.iterator(chunk_size=1000))
    _escribir_filas(ws, filas)